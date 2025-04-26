"""
Main application module for the training form.

This module defines the Flask application and routes.
"""

import os
import csv
import json
import logging
from datetime import datetime
from models import get_db
from io import BytesIO
import functools
from collections import defaultdict

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    send_from_directory,
    url_for,
    flash,
    jsonify,
    abort,
    send_file,
)
from werkzeug.utils import secure_filename
from flask_login import login_user, logout_user, current_user, login_required

from forms import TrainingForm, SearchForm, LoginForm
from models import (
    insert_training_form,
    update_training_form,
    get_all_training_forms,
    get_training_form,
    get_approved_forms_for_export,
    get_user_training_forms,
)
from setup_db import setup_database
from auth import init_auth, authenticate_user, is_admin_email

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Create and configure the Flask application
app = Flask(__name__)

# Load configuration from config.py
app.config.from_pyfile("config.py")

# Set the secret key for CSRF protection
app.secret_key = app.config["SECRET_KEY"]

# Initialize authentication
init_auth(app)

# Make json module available in templates
app.jinja_env.globals["json"] = json

# Set up the database (create tables only if they do not exist)
setup_database(force_recreate=False)


# Function to get a training form by ID - make available to templates
def get_form_by_id(form_id):
    """Get a form by ID - available in templates"""
    return get_training_form(form_id)


# Make helper functions available to templates
app.jinja_env.globals.update(view_form=get_form_by_id)


# Custom Jinja2 filter to convert JSON strings to Python objects
@app.template_filter("from_json")
def from_json(value):
    """Convert a JSON string to a Python object"""
    try:
        return json.loads(value)
    except Exception:
        return []


# Custom Jinja2 filter to remove a parameter from the query string
@app.template_filter("remove_param")
def remove_param(params, param_name):
    """Remove a parameter from a dict of URL parameters"""
    if isinstance(params, dict):
        new_params = params.copy()
        new_params.pop(param_name, None)
        return new_params
    return params


@app.context_processor
def inject_current_year():
    return {"current_year": datetime.now().year}


def is_admin_user(user):
    return hasattr(user, "is_admin") and user.is_admin


def admin_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin_user(current_user):
            abort(403)
        return f(*args, **kwargs)

    return decorated_function


@app.route("/")
def index():
    """Display the training form or redirect to login if not authenticated"""
    if not current_user.is_authenticated:
        return redirect(url_for("login"))
    form = TrainingForm()
    return render_template("index.html", form=form, now=datetime.now())


@app.route("/home")
def home():
    """Display the home page or redirect to login if not authenticated"""
    if not current_user.is_authenticated:
        return redirect(url_for("login"))
    return render_template("home.html", is_admin=is_admin_user(current_user))


@app.route("/login", methods=["GET", "POST"])
def login():
    """Handle user login via LDAP"""
    # Redirect to home if already logged in
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        # If username doesn't include domain, add it
        if "@" not in username and app.config.get("LDAP_DOMAIN"):
            username = f"{username}@{app.config['LDAP_DOMAIN']}"

        # Authenticate user against LDAP
        user = authenticate_user(username, password)

        if user:
            # Log the user in
            login_user(user)
            logging.info(f"User {username} logged in successfully")

            # Redirect to the requested page or the index
            next_page = request.args.get("next")
            if next_page:
                return redirect(next_page)
            return redirect(url_for("index"))

        # If we get here, authentication failed (flash messages set in authenticate_user)
        logging.warning(f"Failed login attempt for {username}")

    return render_template("login.html", form=form)


@app.route("/logout")
@login_required
def logout():
    """Log the user out and redirect to login page"""
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/submit", methods=["GET", "POST"])
@login_required
def submit_form():
    """Process the form submission"""
    form = TrainingForm()

    # Log all form data received
    logging.debug("=== Form Submission Debug ===")
    logging.debug(f"Form data: {request.form}")
    logging.debug(f"Trainees data: {request.form.get('trainees_data')}")

    # Validate the form data
    if form.validate_on_submit():
        try:
            # Get trainees data from form
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                form.trainees_data.data = trainees_data

            # Prepare form data using the form's method
            form_data = form.prepare_form_data()
            logging.debug(f"Prepared form data: {form_data}")

            # Add submitter information
            form_data["submitter"] = current_user.email

            # Validate required fields
            if not form_data.get("training_description"):
                flash("Training Description is required", "error")
                return render_template("index.html", form=form)

            # Insert the form data into the database
            form_id = insert_training_form(form_data)
            logging.debug(f"Form inserted with ID: {form_id}")

            # Create a unique folder for attachments
            unique_folder = os.path.join(app.config["UPLOAD_FOLDER"], f"form_{form_id}")
            os.makedirs(unique_folder, exist_ok=True)

            # Process attachments
            # Use request.files directly as WTForms field might not populate from dynamic inputs
            if "attachments" in request.files:
                new_files = request.files.getlist("attachments")
                descriptions = request.form.getlist("attachment_descriptions[]")
                logging.debug(f"Processing {len(new_files)} new attachments.")

                # Ensure descriptions list matches file list length if necessary
                # Pad descriptions if some files didn't get a description input (shouldn't happen with current JS)
                if len(descriptions) < len(new_files):
                    descriptions.extend([""] * (len(new_files) - len(descriptions)))

                for i, file in enumerate(new_files):
                    # Check if the file object exists and has a filename
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        # Save to the unique folder
                        file_path = os.path.join(unique_folder, filename)
                        logging.debug(f"Saving file {filename} to {file_path}")
                        file.save(file_path)

                        # Get description or use empty string
                        description = descriptions[i] if i < len(descriptions) else ""
                        logging.debug(f"Attachment description: {description}")

                        # Insert attachment record
                        try:
                            conn = get_db()
                            cursor = conn.cursor()
                            cursor.execute(
                                """
                                INSERT INTO attachments (training_id, filename, description)
                                VALUES (?, ?, ?)
                            """,
                                (form_id, filename, description),
                            )
                            conn.commit()
                        except Exception as db_err:
                            logging.error(
                                f"Database error inserting attachment {filename}: {db_err}"
                            )
                        finally:
                            if conn:
                                conn.close()
                    else:
                        logging.debug(f"Skipping empty file input at index {i}.")

            # If no files were submitted via request.files['attachments'], log that.
            elif not request.files.getlist("attachments"):
                logging.debug(
                    "No new files found in request.files for key 'attachments'."
                )

            flash("Form submitted successfully!", "success")
            return redirect(url_for("success"))
        except Exception as e:
            logging.error(f"Error processing form submission: {e}", exc_info=True)
            flash(
                "An error occurred while submitting the form. Please try again.",
                "danger",
            )
            return render_template("index.html", form=form, now=datetime.now())
    else:
        logging.error(f"Form validation errors: {form.errors}")
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {getattr(form, field).label.text}: {error}", "danger")
        return render_template("index.html", form=form, now=datetime.now())


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    """Serve uploaded files with proper content type handling"""
    # Get the directory and filename from the path
    directory = os.path.dirname(filename)
    filename = os.path.basename(filename)

    # Get the file extension
    file_ext = os.path.splitext(filename)[1].lower()

    # Define viewable file types
    viewable_types = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".txt", ".pdf"}

    # Determine if file should be viewed or downloaded
    if file_ext in viewable_types:
        return send_from_directory(
            os.path.join(app.config["UPLOAD_FOLDER"], directory),
            filename,
            as_attachment=False,
        )
    else:
        return send_from_directory(
            os.path.join(app.config["UPLOAD_FOLDER"], directory),
            filename,
            as_attachment=True,
        )


@app.route("/approve/<int:form_id>")
@login_required
@admin_required
def approve_training(form_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE training_forms SET approved = NOT approved WHERE id = ?", (form_id,)
    )
    conn.commit()
    conn.close()

    # Determine redirect target based on referrer
    referrer = request.referrer
    if referrer and url_for("list_forms") in referrer:
        return redirect(referrer)
    else:
        return redirect(url_for("view_form", form_id=form_id))


@app.route("/success")
@login_required
def success():
    """Display success page after form submission"""
    return render_template("success.html", now=datetime.now())


@app.route("/list")
@login_required
def list_forms():
    """Display a list of all training form submissions"""
    form = SearchForm()

    # Get filter parameters from request
    search_term = request.args.get("search", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    training_type = request.args.get("training_type", "")
    sort_by = request.args.get("sort_by", "submission_date")
    sort_order = request.args.get("sort_order", "DESC")
    page = request.args.get("page", 1, type=int)

    # Get forms with filters
    forms, total_count = get_all_training_forms(
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
    )

    # Calculate pagination
    total_pages = (total_count + 9) // 10  # Round up division

    # Create params dictionary for maintaining filters in pagination
    params = {
        "search": search_term,
        "date_from": date_from,
        "date_to": date_to,
        "training_type": training_type,
        "sort_by": sort_by,
        "sort_order": sort_order,
    }

    return render_template(
        "list.html",
        form=form,
        forms=forms,
        total_count=total_count,
        total_pages=total_pages,
        current_page=page,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        sort_by=sort_by,
        sort_order=sort_order,
        now=datetime.now(),
        params=params,
        has_filters=bool(search_term or date_from or date_to or training_type),
        total_forms=total_count,
        is_admin=is_admin_user(current_user),
    )


@app.route("/view/<int:form_id>")
@login_required
def view_form(form_id):
    """Display a single training form submission"""
    form_data = get_training_form(form_id)
    print("DEBUG - Form Data Received:", form_data)
    if not form_data:
        flash("Training form not found", "danger")
        return redirect(url_for("list_forms"))

    # Get attachments
    conn = get_db()
    cursor = conn.cursor()

    # Get attachments records
    cursor.execute("SELECT * FROM attachments WHERE training_id = ?", (form_id,))
    attachments = [dict(row) for row in cursor.fetchall()]

    # Close connection
    conn.close()
    # Parse trainees data from JSON if it exists
    trainees = []
    if form_data.get("trainees_data"):
        try:
            trainees = json.loads(form_data["trainees_data"])
            if (
                isinstance(trainees, list)
                and len(trainees) > 0
                and not isinstance(trainees[0], dict)
            ):
                trainees = [{"email": email} for email in trainees]
        except Exception:
            trainees = []
    # Parse comma-separated trainee emails
    trainee_emails = []
    if form_data.get("trainee_emails"):
        trainee_emails = [
            e.strip() for e in form_data["trainees_data"].split(",") if e.strip()
        ]

    print("Attachments from DB:", attachments)
    return render_template(
        "view.html",
        form=form_data,
        trainees=trainees,
        attachments=attachments,
        trainee_emails=trainee_emails,
        now=datetime.now(),
        is_admin=is_admin_user(current_user),
    )


@app.route("/edit/<int:form_id>", methods=["GET", "POST"])
@login_required
def edit_form(form_id):
    """Edit an existing training form submission"""
    logging.debug(f"Edit form request - Method: {request.method}, Form ID: {form_id}")
    form = TrainingForm()

    if request.method == "GET":
        # Load existing form data
        form_data = get_training_form(form_id)
        if form_data:
            # Basic fields
            form.training_type.data = form_data["training_type"]
            form.trainer_name.data = form_data["trainer_name"]
            form.supplier_name.data = form_data["supplier_name"]
            form.location_type.data = form_data["location_type"]
            form.location_details.data = form_data["location_details"]

            # Date fields
            try:
                form.start_date.data = datetime.strptime(
                    form_data["start_date"], "%Y-%m-%d"
                )
                form.end_date.data = datetime.strptime(
                    form_data["end_date"], "%Y-%m-%d"
                )
            except (ValueError, KeyError) as e:
                logging.error(f"Error parsing dates: {str(e)}")
                flash("Error loading date information")
                return redirect(url_for("list_forms"))

            # Numeric fields
            form.trainer_hours.data = form_data["trainer_hours"]
            form.training_description.data = form_data.get("training_description", "")

            # Expense fields
            form.travel_cost.data = form_data.get("travel_cost", 0)
            form.food_cost.data = form_data.get("food_cost", 0)
            form.materials_cost.data = form_data.get("materials_cost", 0)
            form.other_cost.data = form_data.get("other_cost", 0)
            form.other_expense_description.data = form_data.get(
                "other_expense_description", ""
            )
            form.concur_claim.data = form_data.get("concur_claim", "")
            form.trainee_hours.data = form_data.get("trainee_hours", 0)

            # Load trainees data
            if form_data.get("trainees_data"):
                form.trainees_data.data = form_data.get("trainees_data")

    if form.validate_on_submit():
        try:
            # Get trainees data from form
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                form.trainees_data.data = trainees_data

            # Get form data
            form_data = form.prepare_form_data()
            logging.debug(f"Prepared form data: {form_data}")

            # Get the existing form to preserve the submitter
            existing_form = get_training_form(form_id)
            if existing_form and existing_form.get("submitter"):
                form_data["submitter"] = existing_form["submitter"]
            else:
                form_data["submitter"] = current_user.email

            # Validate required fields
            if not form_data.get("training_description"):
                flash("Training Description is required", "error")
                return render_template(
                    "index.html", form=form, edit_mode=True, form_id=form_id
                )

            # Update form data in database
            update_training_form(form_id, form_data)
            logging.debug(f"Form {form_id} core data updated.")

            # --- Handle Attachments ---
            unique_folder = os.path.join(app.config["UPLOAD_FOLDER"], f"form_{form_id}")
            os.makedirs(unique_folder, exist_ok=True)
            logging.debug(f"Ensured attachment directory exists: {unique_folder}")

            # Process NEW attachments
            # Use request.files directly as WTForms field might not populate from dynamic inputs
            if "attachments" in request.files:
                new_files = request.files.getlist("attachments")
                descriptions = request.form.getlist(
                    "attachment_descriptions[]"
                )  # Descriptions for NEW files
                logging.debug(
                    f"Processing {len(new_files)} new attachments for form {form_id}."
                )

                # Pad descriptions if necessary (belt-and-suspenders)
                if len(descriptions) < len(new_files):
                    descriptions.extend([""] * (len(new_files) - len(descriptions)))

                for i, file in enumerate(new_files):
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        # SAVE TO FORM-SPECIFIC FOLDER
                        file_path = os.path.join(unique_folder, filename)
                        logging.debug(f"Saving new file {filename} to {file_path}")
                        file.save(file_path)

                        description = descriptions[i] if i < len(descriptions) else ""
                        logging.debug(f"New attachment description: {description}")

                        # Insert new attachment record
                        try:
                            conn = get_db()
                            cursor = conn.cursor()
                            cursor.execute(
                                """
                                INSERT INTO attachments (training_id, filename, description)
                                VALUES (?, ?, ?)
                            """,
                                (form_id, filename, description),
                            )
                            conn.commit()
                        except Exception as db_err:
                            logging.error(
                                f"Database error inserting new attachment {filename}: {db_err}"
                            )
                        finally:
                            if conn:
                                conn.close()
                    else:
                        logging.debug(f"Skipping empty new file input at index {i}.")
            else:
                logging.debug(
                    "No new files found in request.files for key 'attachments'."
                )

            # Handle attachment DELETIONS
            delete_attachments = request.form.getlist("delete_attachments[]")
            if delete_attachments:
                logging.debug(
                    f"Processing deletions for attachment IDs: {delete_attachments}"
                )
                # TODO: Optionally delete the actual files from the filesystem
                try:
                    conn = get_db()
                    cursor = conn.cursor()
                    # Placeholders for safe query
                    placeholders = ", ".join("?" * len(delete_attachments))
                    query = f"DELETE FROM attachments WHERE id IN ({placeholders}) AND training_id = ?"
                    params = delete_attachments + [form_id]  # Add form_id for security
                    cursor.execute(query, params)
                    conn.commit()
                    logging.info(
                        f"Deleted {cursor.rowcount} attachment records for form {form_id}."
                    )
                except Exception as db_err:
                    logging.error(f"Database error deleting attachments: {db_err}")
                finally:
                    if conn:
                        conn.close()

            # Handle attachment description UPDATES for existing files
            update_descriptions = request.form.getlist(
                "update_attachment_descriptions[]"
            )
            if update_descriptions:
                logging.debug(f"Processing description updates: {update_descriptions}")
                try:
                    conn = get_db()
                    cursor = conn.cursor()
                    for desc_json in update_descriptions:
                        try:
                            desc_data = json.loads(desc_json)
                            att_id = desc_data.get("id")
                            new_desc = desc_data.get("description", "")
                            if att_id:
                                cursor.execute(
                                    """
                                    UPDATE attachments SET description = ? 
                                    WHERE id = ? AND training_id = ?
                                    """,
                                    (new_desc, att_id, form_id),  # Add form_id check
                                )
                            else:
                                logging.warning(
                                    f"Skipping description update due to missing ID in JSON: {desc_json}"
                                )
                        except json.JSONDecodeError:
                            logging.error(
                                f"Invalid JSON in attachment description update: {desc_json}"
                            )
                        except KeyError:
                            logging.error(
                                f"Missing 'id' or 'description' key in JSON: {desc_json}"
                            )
                    conn.commit()
                except Exception as db_err:
                    logging.error(
                        f"Database error updating attachment descriptions: {db_err}"
                    )
                finally:
                    if conn:
                        conn.close()

            flash("Form updated successfully!", "success")
            return redirect(url_for("view_form", form_id=form_id))

        except Exception as e:
            logging.error(f"Error updating form: {str(e)}")
            flash(
                "An error occurred while updating the form. Please try again.", "danger"
            )

    # Load existing attachments to display in form
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM attachments WHERE training_id = ?", (form_id,))
    existing_attachments = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return render_template(
        "index.html",
        form=form,
        edit_mode=True,
        form_id=form_id,
        existing_attachments=existing_attachments,
    )


@app.route("/api/employees")
@login_required
def get_employees():
    """API endpoint to get employees from the CSV file"""
    try:
        employees = []
        csv_path = os.path.join("attached_assets", "EmployeeListFirstLastDept.csv")

        if not os.path.exists(csv_path):
            logging.error(f"Employee list CSV not found at {csv_path}")
            return jsonify([])

        logging.info(f"Loading employee data from {csv_path}")

        with open(csv_path, "r", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Get first name, last name, email, and department
                first_name = row.get("FirstName", "").strip()
                last_name = row.get("LastName", "").strip()
                email = row.get("UserPrincipalName", "").strip()
                department = row.get("Department", "").strip()

                # Create display name from first and last name
                if first_name and last_name and email:
                    display_name = f"{first_name} {last_name}"
                    employees.append(
                        {
                            "displayName": display_name,
                            "email": email,
                            "name": display_name,
                            "department": department,
                            "firstName": first_name,
                            "lastName": last_name,
                        }
                    )

        logging.info(f"Loaded {len(employees)} employees")
        return jsonify(employees)
    except Exception as e:
        logging.error(f"Error loading employee data: {str(e)}")
        return jsonify([])


@app.route("/api/export_claim5_options")
@login_required
def export_claim5_options():
    if not is_admin_user(current_user):
        return jsonify({"error": "Unauthorized"}), 403
    from models import get_approved_forms_for_export
    import datetime

    forms = get_approved_forms_for_export()
    if not forms:
        return jsonify({"quarters": [], "min_date": None, "max_date": None})
    created_dates = [
        f.get("created_at") or f.get("submission_date") or f.get("start_date")
        for f in forms
        if f.get("created_at") or f.get("submission_date") or f.get("start_date")
    ]
    created_dates = [
        datetime.datetime.fromisoformat(str(d)[:10]) for d in created_dates
    ]
    min_date = min(created_dates).date().isoformat()
    max_date = max(created_dates).date().isoformat()

    def get_quarter(dt):
        q = (dt.month - 1) // 3 + 1
        return f"Q{q} {dt.year}"

    quarters = sorted(
        {get_quarter(dt) for dt in created_dates}, key=lambda x: (int(x[1]), int(x[3:]))
    )
    return jsonify({"quarters": quarters, "min_date": min_date, "max_date": max_date})


@app.route("/export_claim5", methods=["GET", "POST"])
@login_required
def export_claim5():
    if not is_admin_user(current_user):
        flash("Unauthorized", "danger")
        return redirect(url_for("list_forms"))
    import json

    if request.method == "POST":
        data = request.get_json()
        selected_quarters = data.get("quarters", [])
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        from models import get_approved_forms_for_export
        import datetime

        forms = get_approved_forms_for_export()
        filtered_forms = []

        def get_quarter(dt):
            q = (dt.month - 1) // 3 + 1
            return f"Q{q} {dt.year}"

        for f in forms:
            created = (
                f.get("created_at") or f.get("submission_date") or f.get("start_date")
            )
            if not created:
                continue
            dt = datetime.datetime.fromisoformat(str(created)[:10])
            quarter = get_quarter(dt)
            if selected_quarters and quarter in selected_quarters:
                filtered_forms.append(f)
            elif start_date and end_date:
                if start_date <= dt.date().isoformat() <= end_date:
                    filtered_forms.append(f)
        approved_forms = filtered_forms
    else:
        from models import get_approved_forms_for_export

        approved_forms = get_approved_forms_for_export()

    try:
        # Path to the template Excel file
        template_path = os.path.join(
            "attached_assets", "Claim-Form-5-revised-Training.xlsx"
        )

        # Load the template using openpyxl
        from openpyxl import load_workbook

        if not os.path.exists(template_path):
            flash("Template file not found.", "danger")
            return redirect(url_for("list_forms"))

        wb = load_workbook(template_path)
        ws = wb.active  # Assuming the template has only one sheet

        # Column headers are in row 8, data starts from row 9
        start_row = 9
        current_row = start_row

        # Process each training form
        for form in approved_forms:
            # Get trainees data for this form
            trainees = []
            if form.get("trainees_data"):
                try:
                    trainees_data = json.loads(form["trainees_data"])
                    if isinstance(trainees_data, list):
                        # Handle different possible formats of trainees_data
                        if trainees_data and isinstance(trainees_data[0], dict):
                            trainees = trainees_data
                        else:
                            # Simple list of emails, convert to dict format
                            trainees = [
                                {"email": email, "name": email}
                                for email in trainees_data
                            ]
                except json.JSONDecodeError:
                    logging.error(f"Error parsing trainees_data for form {form['id']}")
                    trainees = []

            # If no trainees found, add a placeholder row
            if not trainees:
                trainees = [{"name": "Unknown", "email": ""}]

            # Determine location value based on location_type
            if form.get("location_type") == "Onsite":
                location = "Onsite (Stryker Limerick)"
            else:
                location = form.get("location_details", "")

            # Process each trainee
            for i, trainee in enumerate(trainees):
                trainee_name = trainee.get(
                    "name", trainee.get("email", "Unknown Trainee")
                )

                # Insert new rows if we're exceeding the template's initial capacity
                if current_row > 23:  # 23 is the last empty row in the template
                    ws.insert_rows(current_row)

                # Fill the row with data according to requirements
                ws.cell(row=current_row, column=1).value = (
                    trainee_name  # Names of Trainees
                )
                ws.cell(row=current_row, column=2).value = location  # Location
                ws.cell(row=current_row, column=3).value = ""  # Weekly Wage (blank)
                ws.cell(row=current_row, column=4).value = form.get(
                    "trainee_hours", ""
                )  # Nr of Weeks/days/hours
                ws.cell(row=current_row, column=5).value = ""

                # Only fill these fields for the first trainee of each form
                if i == 0:
                    ws.cell(row=current_row, column=6).value = form.get(
                        "trainer_name", ""
                    )  # Name of trainer (moved from col 5)
                    ws.cell(row=current_row, column=7).value = (
                        ""  # Nr of Weeks/days/hours (blank) (moved from col 6)
                    )
                    ws.cell(row=current_row, column=8).value = (
                        location  # Location (same as trainee location) (moved from col 7)
                    )
                    ws.cell(row=current_row, column=9).value = (
                        ""  # Salary (blank) (moved from col 8)
                    )
                    ws.cell(row=current_row, column=10).value = form.get(
                        "supplier_name", ""
                    )  # Supplier/Name (moved from col 9)
                    ws.cell(row=current_row, column=11).value = form.get(
                        "travel_cost", ""
                    )  # Travel (moved from col 10)
                    ws.cell(row=current_row, column=12).value = form.get(
                        "food_cost", ""
                    )  # Subsistence (moved from col 11)
                    ws.cell(row=current_row, column=13).value = form.get(
                        "materials_cost", ""
                    )  # External Trainer / Course Costs (moved from col 12)
                    ws.cell(row=current_row, column=14).value = form.get(
                        "materials_cost", ""
                    )  # Materials (moved from col 13)

                current_row += 1

            # Add an empty row between different training forms
            if current_row <= 23 or len(approved_forms) > 1:
                ws.insert_rows(current_row)
                current_row += 1

        # Create an in-memory file to store the Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logging.info(
            f"Exported {len(approved_forms)} approved forms to Excel template."
        )
        # Send the file to the user
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="claim5_export.xlsx",
        )

    except Exception as e:
        logging.error(f"Error exporting data to Excel template: {e}", exc_info=True)
        flash("An error occurred during the export process.", "danger")
        return redirect(url_for("list_forms"))


@app.errorhandler(500)
def internal_error(error):
    """Handle internal server errors"""
    flash("An internal server error occurred. Please try again later.", "danger")
    return render_template("index.html", form=TrainingForm(), now=datetime.now()), 500


@app.errorhandler(403)
def forbidden(e):
    # Flash a message for 403 errors and redirect to the previous page
    flash("You do not have permission to perform this action.", "danger")
    return redirect(request.referrer or url_for("index")), 403


# Add user information to layout template context
@app.context_processor
def inject_user():
    """Add user information to template context"""
    user_info = {}
    if current_user.is_authenticated:
        user_info = {
            "username": current_user.username,
            "display_name": current_user.display_name or current_user.username,
            "email": current_user.email,
            "first_name": current_user.first_name,
            "last_name": current_user.last_name,
            "is_admin": getattr(
                current_user, "is_admin", False
            ),  # Get admin status if available
        }
    return {"user_info": user_info}


@app.route("/my_submissions")
@login_required
def my_submissions():
    """Display a list of the current user's training form submissions"""
    form = SearchForm()
    search_term = request.args.get("search", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    training_type = request.args.get("training_type", "")
    sort_by = request.args.get("sort_by", "submission_date")
    sort_order = request.args.get("sort_order", "DESC")
    page = request.args.get("page", 1, type=int)

    forms, total_count = get_user_training_forms(
        current_user.email,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
    )
    total_pages = (total_count + 9) // 10
    params = {
        "search": search_term,
        "date_from": date_from,
        "date_to": date_to,
        "training_type": training_type,
        "sort_by": sort_by,
        "sort_order": sort_order,
    }
    return render_template(
        "list.html",
        form=form,
        forms=forms,
        total_count=total_count,
        total_pages=total_pages,
        current_page=page,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        sort_by=sort_by,
        sort_order=sort_order,
        now=datetime.now(),
        params=params,
        has_filters=bool(search_term or date_from or date_to or training_type),
        total_forms=total_count,
        my_submissions=True,
        is_admin=is_admin_user(current_user),
    )


@app.route("/new")
@login_required
def new_form():
    """Display the training form"""
    form = TrainingForm()
    return render_template("index.html", form=form, now=datetime.now())


@app.route("/leaderboard")
@login_required
def leaderboard():
    """Display the leaderboard of trainers by total training hours"""


    # Get all approved forms
    forms = get_approved_forms_for_export()
    trainer_hours = defaultdict(float)

    for form in forms:
        trainer_name = form.get("trainer_name") or "Unknown"
        try:
            trainees_data = json.loads(form.get("trainees_data") or "[]")
            if isinstance(trainees_data, list):
                if trainees_data and isinstance(trainees_data[0], dict):
                    num_trainees = len(trainees_data)
                else:
                    num_trainees = len(trainees_data)
            else:
                num_trainees = 0
        except Exception:
            num_trainees = 0
        trainer_hours_val = float(form.get("trainer_hours") or 0)
        trainee_hours_val = float(form.get("trainee_hours") or 0)
        total_hours = trainer_hours_val + (trainee_hours_val * num_trainees)
        trainer_hours[trainer_name] += total_hours

    # Sort trainers by total hours descending
    leaderboard_data = sorted(trainer_hours.items(), key=lambda x: x[1], reverse=True)
    names = [x[0] for x in leaderboard_data]
    hours = [x[1] for x in leaderboard_data]
    return render_template("leaderboard.html", names=names, hours=hours)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=app.config["DEBUG"])
