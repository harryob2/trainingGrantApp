"""
Main application module for the training form.

This module defines the Flask application and routes.
"""

import os
import csv
import json
import logging
from datetime import datetime
from io import BytesIO
import functools
from collections import defaultdict
from openpyxl import load_workbook
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    send_from_directory,
    url_for,
    flash,
    jsonify,
    abort,
    send_file,
    render_template_string,
)
from werkzeug.utils import secure_filename
from flask_login import login_user, logout_user, current_user, login_required

from forms import TrainingForm, SearchForm, LoginForm
from models import (
    insert_training_form,
    update_training_form,
    get_all_training_forms,
    get_training_form,
    get_approved_forms_for_export,
    get_user_training_forms,
    add_admin,
    db_session,
    Admin,
    Attachment,
    soft_delete_training_form,
    recover_training_form,
)
from setup_db import setup_database
from auth import init_auth, authenticate_user, is_admin_email
from lookups import get_lookup_data
from utils import get_quarter
from email_utils import init_mail, send_form_submission_notification

# Import our new logging configuration
from logging_config import setup_logging, get_logger

# Create and configure the Flask application
app = Flask(__name__)

# Load configuration from config.py
app.config.from_pyfile("config.py")

# Set the secret key for CSRF protection
app.secret_key = app.config["SECRET_KEY"]

# Initialize centralized logging
setup_logging(app)

# Get logger for this module
logger = get_logger(__name__)

# Initialize authentication
init_auth(app)

# Initialize email
init_mail(app)

# Make json module available in templates
app.jinja_env.globals["json"] = json

# Set up the database (create tables only if they do not exist)
setup_database(force_recreate=False)

# Ensure upload folder exists
upload_folder = app.config["UPLOAD_FOLDER"]
try:
    os.makedirs(upload_folder, exist_ok=True)
    logger.info(f"Upload folder configured: {upload_folder}")
except Exception as e:
    logger.error(f"Failed to create upload folder {upload_folder}: {e}")
    raise

# Function to get a training form by ID - make available to templates
def get_form_by_id(form_id):
    """Get a form by ID - available in templates"""
    return get_training_form(form_id)


# Make helper functions available to templates
app.jinja_env.globals.update(view_form=get_form_by_id)


# Custom Jinja2 filter to convert JSON strings to Python objects
@app.template_filter("from_json")
def from_json(value):
    """Convert a JSON string to a Python object"""
    try:
        return json.loads(value)
    except Exception:
        return []


# Custom Jinja2 filter to remove a parameter from the query string
@app.template_filter("remove_param")
def remove_param(params, param_name):
    """Remove a parameter from a dict of URL parameters"""
    if isinstance(params, dict):
        new_params = params.copy()
        new_params.pop(param_name, None)
        return new_params
    return params


@app.context_processor
def inject_current_year():
    return {"current_year": datetime.now().year}


def is_admin_user(user):
    return hasattr(user, "email") and is_admin_email(user.email)


def admin_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin_user(current_user):
            abort(403)
        return f(*args, **kwargs)

    return decorated_function


@app.route("/")
def index():
    """Display the training form or redirect to login if not authenticated"""
    if not current_user.is_authenticated:
        return redirect(url_for("login"))
    return render_template("home.html", is_admin=is_admin_user(current_user))


@app.route("/login", methods=["GET", "POST"])
def login():
    """Handle user login via LDAP"""
    # Redirect to home if already logged in
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        # If username doesn't include domain, add it
        if "@" not in username and app.config.get("LDAP_DOMAIN"):
            username = f"{username}@{app.config['LDAP_DOMAIN']}"

        # Authenticate user against LDAP
        user = authenticate_user(username, password, app.config)

        if user:
            # Log the user in
            login_user(user)
            logger.info(
                "User login successful",
                extra={
                    "user": username,
                    "ip_address": request.remote_addr,
                    "user_agent": str(request.user_agent)
                }
            )

            # Redirect to the requested page or the index
            next_page = request.args.get("next")
            if next_page:
                return redirect(next_page)
            return redirect(url_for("index"))

        # If we get here, authentication failed (flash messages set in authenticate_user)
        logger.warning(
            "Failed login attempt",
            extra={
                "user": username,
                "ip_address": request.remote_addr,
                "user_agent": str(request.user_agent)
            }
        )

    return render_template("login.html", form=form)


@app.route("/manage_admins", methods=["GET", "POST"])
@login_required
@admin_required
def manage_admins():
    message = None
    admins = []
    if request.method == "POST":
        if "add_admin" in request.form:
            email = request.form["email"].strip().lower()
            first_name = request.form["first_name"].strip()
            last_name = request.form["last_name"].strip()
            success = add_admin(
                {"email": email, "first_name": first_name, "last_name": last_name}
            )
            if not success:
                flash("Admin already exists.", "warning")
            else:
                flash("Admin added.", "success")
        elif "remove_admin" in request.form:
            email = request.form["remove_admin"].strip().lower()
            with db_session() as session:
                admin = session.query(Admin).filter_by(email=email).first()
                if admin:
                    session.delete(admin)
                    flash("Admin removed.", "success")
    
    with db_session() as session:
        admins = session.query(Admin).all()
        admins = [
            dict(
                email=a.email, 
                first_name=a.first_name, 
                last_name=a.last_name,
                receive_emails=a.receive_emails
            )
            for a in admins
        ]
    return render_template("manage_admins.html", admins=admins)


@app.route("/update_admin_email_preference", methods=["POST"])
@login_required
@admin_required
def update_admin_email_preference_route():
    """HTMX endpoint to update admin email preference"""
    email = request.form.get("email")
    receive_emails = request.form.get("receive_emails") == "on"
    
    if not email:
        return "Error: Missing email", 400
    
    from models import update_admin_email_preference
    success = update_admin_email_preference(email, receive_emails)
    
    if success:
        if receive_emails:
            return '<span class="text-success small">✓ Enabled</span>'
        else:
            return '<span class="text-danger small">✗ Disabled</span>'
    else:
        return '<span class="text-danger small">Error updating</span>'


@app.route("/logout")
@login_required
def logout():
    """Log the user out and redirect to login page"""
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/submit", methods=["GET", "POST"])
@login_required
def submit_form():
    """Process the form submission"""
    form = TrainingForm()

    if form.validate_on_submit():
        try:
            # Get trainees data from form
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                form.trainees_data.data = trainees_data

            # Prepare form data using the form's method
            form_data = form.prepare_form_data()

            # Add submitter information
            form_data["submitter"] = current_user.email

            # Validate required fields
            if not form_data.get("training_description"):
                flash("Training Description is required", "error")
                return render_template("index.html", form=form)

            # Insert the form data into the database
            form_id = insert_training_form(form_data)
            
            # Log form submission
            logger.info(
                "Training form submitted",
                extra={
                    "form_id": form_id,
                    "submitter": current_user.email,
                    "training_type": form_data.get("training_type"),
                    "training_name": form_data.get("training_name"),
                    "start_date": str(form_data.get("start_date")),
                    "end_date": str(form_data.get("end_date"))
                }
            )

            # Create a unique folder for attachments
            unique_folder = os.path.join(app.config["UPLOAD_FOLDER"], f"form_{form_id}")
            os.makedirs(unique_folder, exist_ok=True)

            # Process attachments using SQLAlchemy ORM
            if "attachments" in request.files:
                new_files = request.files.getlist("attachments")
                descriptions = request.form.getlist("attachment_descriptions[]")

                if len(descriptions) < len(new_files):
                    descriptions.extend([""] * (len(new_files) - len(descriptions)))

                with db_session() as session:
                    for i, file in enumerate(new_files):
                        if file and file.filename:
                            filename = secure_filename(file.filename)
                            file_path = os.path.join(unique_folder, filename)
                            file.save(file_path)
                            description = (
                                descriptions[i] if i < len(descriptions) else ""
                            )
                            session.add(
                                Attachment(
                                    form_id=form_id,
                                    filename=filename,
                                    description=description,
                                )
                            )

            # Process travel expenses
            travel_expenses_data = request.form.get("travel_expenses_data")
            if travel_expenses_data:
                try:
                    from models import insert_travel_expenses
                    travel_expenses = json.loads(travel_expenses_data)
                    if travel_expenses and isinstance(travel_expenses, list):
                        insert_travel_expenses(form_id, travel_expenses)
                        logger.info(f"Inserted {len(travel_expenses)} travel expenses for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logger.error(
                        "Error processing travel expenses",
                        extra={
                            "form_id": form_id,
                            "error": str(e),
                            "submitter": current_user.email
                        },
                        exc_info=True
                    )
                    # Don't fail the form submission for travel expense errors
                    flash("Warning: There was an issue processing travel expenses, but the form was submitted successfully.", "warning")

            # Process material expenses
            material_expenses_data = request.form.get("material_expenses_data")
            if material_expenses_data:
                try:
                    from models import insert_material_expenses
                    material_expenses = json.loads(material_expenses_data)
                    if material_expenses and isinstance(material_expenses, list):
                        insert_material_expenses(form_id, material_expenses)
                        logger.info(f"Inserted {len(material_expenses)} material expenses for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logger.error(
                        "Error processing material expenses",
                        extra={
                            "form_id": form_id,
                            "error": str(e),
                            "submitter": current_user.email
                        },
                        exc_info=True
                    )
                    # Don't fail the form submission for material expense errors
                    flash("Warning: There was an issue processing material expenses, but the form was submitted successfully.", "warning")

            # Process trainees using the new table structure
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                try:
                    from models import insert_trainees
                    trainees = json.loads(trainees_data)
                    if trainees and isinstance(trainees, list):
                        insert_trainees(form_id, trainees)
                        logger.info(f"Inserted {len(trainees)} trainees for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logger.error(
                        "Error processing trainees",
                        extra={
                            "form_id": form_id,
                            "error": str(e),
                            "submitter": current_user.email
                        },
                        exc_info=True
                    )
                    # Don't fail the form submission for trainee errors
                    flash("Warning: There was an issue processing trainees, but the form was submitted successfully.", "warning")

            # Send email notification
            try:
                send_form_submission_notification(form_id, form_data, current_user.email)
            except Exception as e:
                logger.error(
                    "Failed to send email notification",
                    extra={
                        "form_id": form_id,
                        "error": str(e),
                        "submitter": current_user.email
                    },
                    exc_info=True
                )
                # Don't fail the form submission if email fails

            flash("Form submitted successfully!", "success")
            return redirect(url_for("success"))
        except Exception as e:
            logger.error(
                "Error processing form submission",
                extra={
                    "submitter": current_user.email,
                    "error": str(e)
                },
                exc_info=True
            )
            flash(
                "An error occurred while submitting the form. Please try again.",
                "danger",
            )
            return render_template("index.html", form=form, now=datetime.now())
    else:
        # Flash form validation errors
        for field, errors in form.errors.items():
            field_obj = getattr(form, field, None)
            field_label = field_obj.label.text if field_obj and hasattr(field_obj, 'label') else field
            for error in errors:
                flash(f"Error in {field_label}: {error}", "danger")
        
        return render_template("index.html", form=form, now=datetime.now())


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    """Serve uploaded files with proper content type handling"""
    # Get the directory and filename from the path
    directory = os.path.dirname(filename)
    filename = os.path.basename(filename)

    # Get the file extension
    file_ext = os.path.splitext(filename)[1].lower()

    # Define viewable file types
    viewable_types = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".txt", ".pdf"}

    # Determine if file should be viewed or downloaded
    if file_ext in viewable_types:
        return send_from_directory(
            os.path.join(app.config["UPLOAD_FOLDER"], directory),
            filename,
            as_attachment=False,
        )
    else:
        return send_from_directory(
            os.path.join(app.config["UPLOAD_FOLDER"], directory),
            filename,
            as_attachment=True,
        )


@app.route("/approve/<int:form_id>")
@login_required
@admin_required
def approve_training(form_id):
    from models import TrainingForm

    with db_session() as session:
        form = session.query(TrainingForm).filter_by(id=form_id).first()
        if form:
            was_approved = bool(form.approved)
            form.approved = not was_approved
            session.flush()
            
            # Log approval action
            logger.info(
                "Form approval status changed",
                extra={
                    "form_id": form_id,
                    "admin": current_user.email,
                    "action": "approved" if not was_approved else "unapproved",
                    "new_status": not was_approved
                }
            )

    # If htmx request for row update in list
    if request.args.get("row") == "1":
        from models import get_training_form

        form_data = get_training_form(form_id)
        is_admin = is_admin_user(current_user)
        return render_template("_form_row.html", form=form_data, is_admin=is_admin)

    # If htmx request for view page button
    if request.args.get("view") == "1":
        from models import get_training_form

        form_data = get_training_form(form_id)
        is_admin = is_admin_user(current_user)
        return render_template(
            "_approve_btn_view.html", form=form_data, is_admin=is_admin
        )

    # Determine redirect target based on referrer
    referrer = request.referrer
    if referrer and url_for("list_forms") in referrer:
        return redirect(referrer)
    else:
        return redirect(url_for("view_form", form_id=form_id))


@app.route("/delete/<int:form_id>", methods=["POST"])
@login_required
def delete_training_form(form_id):
    """Soft delete a training form"""
    # Get the form to check permissions (including deleted forms for recovery)
    form_data = get_training_form(form_id, include_deleted=True)
    if not form_data:
        flash("Training form not found", "danger")
        return redirect(url_for("list_forms"))
    
    # Check if user is admin or the submitter
    if not (is_admin_user(current_user) or form_data.get("submitter") == current_user.email):
        flash("You don't have permission to delete this form", "danger")
        return redirect(url_for("view_form", form_id=form_id))
    
    # Perform soft delete
    if soft_delete_training_form(form_id):
        logger.warning(
            "Training form deleted",
            extra={
                "form_id": form_id,
                "deleted_by": current_user.email,
                "is_admin": is_admin_user(current_user),
                "original_submitter": form_data.get("submitter")
            }
        )
        flash("Training form has been deleted successfully", "success")
        # Redirect to appropriate list based on user role
        if is_admin_user(current_user):
            return redirect(url_for("list_forms"))
        else:
            return redirect(url_for("my_submissions"))
    else:
        logger.error(
            "Error deleting training form",
            extra={
                "form_id": form_id,
                "attempted_by": current_user.email
            }
        )
        flash("Error deleting training form", "danger")
        return redirect(url_for("view_form", form_id=form_id))


@app.route("/recover/<int:form_id>", methods=["POST"])
@login_required
def recover_training_form_route(form_id):
    """Recover a soft deleted training form"""
    # Get the form to check permissions (including deleted forms)
    form_data = get_training_form(form_id, include_deleted=True)
    if not form_data:
        flash("Training form not found", "danger")
        return redirect(url_for("list_forms"))
    
    # Check if user is admin or the submitter
    if not (is_admin_user(current_user) or form_data.get("submitter") == current_user.email):
        flash("You don't have permission to recover this form", "danger")
        return redirect(url_for("view_form", form_id=form_id))
    
    # Perform recovery
    if recover_training_form(form_id):
        logger.info(
            "Training form recovered",
            extra={
                "form_id": form_id,
                "recovered_by": current_user.email,
                "is_admin": is_admin_user(current_user),
                "original_submitter": form_data.get("submitter")
            }
        )
        flash("Training form has been recovered successfully", "success")
        # Redirect to appropriate list based on user role
        if is_admin_user(current_user):
            return redirect(url_for("list_forms"))
        else:
            return redirect(url_for("my_submissions"))
    else:
        logger.error(
            "Error recovering training form",
            extra={
                "form_id": form_id,
                "attempted_by": current_user.email
            }
        )
        flash("Error recovering training form", "danger")
        return redirect(url_for("view_form", form_id=form_id))


@app.route("/success")
@login_required
def success():
    """Display success page after form submission"""
    return render_template("success.html", now=datetime.now())


@app.route("/list")
@login_required
def list_forms():
    """Display a list of all training form submissions"""
    form = SearchForm()

    # Get filter parameters from request
    search_term = request.args.get("search", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    training_type = request.args.get("training_type", "")
    approval_status = request.args.get("approval_status", "")
    delete_status = request.args.get("delete_status", "")
    sort_by = request.args.get("sort_by", "submission_date")
    sort_order = request.args.get("sort_order", "DESC")
    page = request.args.get("page", 1, type=int)

    # Populate form fields with current filter values
    form.search.data = search_term
    if date_from:
        try:
            form.date_from.data = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_to:
        try:
            form.date_to.data = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass
    form.training_type.data = training_type
    form.approval_status.data = approval_status
    form.delete_status.data = delete_status
    form.sort_by.data = sort_by
    form.sort_order.data = sort_order

    # Get forms with filters
    forms, total_count = get_all_training_forms(
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        approval_status=approval_status,
        delete_status=delete_status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
    )

    # Calculate pagination
    total_pages = (total_count + 9) // 10  # Round up division

    # Create params dictionary for maintaining filters in pagination
    params = {
        "search": search_term,
        "date_from": date_from,
        "date_to": date_to,
        "training_type": training_type,
        "approval_status": approval_status,
        "delete_status": delete_status,
        "sort_by": sort_by,
        "sort_order": sort_order,
    }

    return render_template(
        "list.html",
        form=form,
        forms=forms,
        total_count=total_count,
        total_pages=total_pages,
        current_page=page,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        approval_status=approval_status,
        delete_status=delete_status,
        sort_by=sort_by,
        sort_order=sort_order,
        now=datetime.now(),
        params=params,
        has_filters=bool(search_term or date_from or date_to or training_type or approval_status or delete_status),
        total_forms=total_count,
        is_admin=is_admin_user(current_user),
    )


@app.route("/view/<int:form_id>")
@login_required
def view_form(form_id):
    """Display a single training form submission"""
    form_data = get_training_form(form_id, include_deleted=True)
    if not form_data:
        flash("Training form not found", "danger")
        return redirect(url_for("list_forms"))

    # Get attachments
    with db_session() as session:
        attachments = session.query(Attachment).filter_by(form_id=form_id).all()
        attachments = [
            {
                "id": a.id,
                "form_id": a.form_id,
                "filename": a.filename,
                "description": a.description,
            }
            for a in attachments
        ]
    
    # Get trainees from the new table structure
    trainees = []
    try:
        from models import get_trainees
        trainees = get_trainees(form_id)
    except Exception as e:
        logging.error(f"Error loading trainees for view: {e}")

    # Get travel expenses
    travel_expenses = []
    try:
        from models import get_travel_expenses
        travel_expenses = get_travel_expenses(form_id)
    except Exception as e:
        logging.error(f"Error loading travel expenses for view: {e}")

    # Get material expenses
    material_expenses = []
    try:
        from models import get_material_expenses
        material_expenses = get_material_expenses(form_id)
    except Exception as e:
        logging.error(f"Error loading material expenses for view: {e}")

    return render_template(
        "view.html",
        form=form_data,
        trainees=trainees,
        attachments=attachments,
        travel_expenses=travel_expenses,
        material_expenses=material_expenses,
        now=datetime.now(),
        is_admin=is_admin_user(current_user),
    )


@app.route("/edit/<int:form_id>", methods=["GET", "POST"])
@login_required
def edit_form(form_id):
    """Edit an existing training form submission"""
    form = TrainingForm()

    if request.method == "GET":
        # Load existing form data
        form_data = get_training_form(form_id)
        if form_data:
            # Basic fields
            form.training_type.data = form_data["training_type"]
            form.training_name.data = form_data.get("training_name", "")  # Load training name
            form.trainer_name.data = form_data["trainer_name"]
            form.trainer_email.data = form_data.get("trainer_email", "")  # Load trainer email
            form.trainer_department.data = form_data.get("trainer_department", "")  # Load trainer department
            form.supplier_name.data = form_data["supplier_name"]
            form.location_type.data = form_data["location_type"]
            form.location_details.data = form_data["location_details"]

            # Date fields
            try:
                form.start_date.data = datetime.strptime(
                    form_data["start_date"], "%Y-%m-%d"
                )
                form.end_date.data = datetime.strptime(
                    form_data["end_date"], "%Y-%m-%d"
                )
            except (ValueError, KeyError) as e:
                logging.error(f"Error parsing dates: {str(e)}")
                flash("Error loading date information")
                return redirect(url_for("list_forms"))

            # Numeric fields
            form.training_hours.data = form_data["training_hours"]
            form.course_cost.data = form_data.get("course_cost", 0)
            form.invoice_number.data = form_data.get("invoice_number", "")
            form.training_description.data = form_data.get("training_description", "")
            form.notes.data = form_data.get("notes", "")
            form.ida_class.data = form_data.get("ida_class", "")

            # Expense fields
            form.concur_claim.data = form_data.get("concur_claim", "")

            # Load trainees data from the new table structure
            try:
                from models import get_trainees
                existing_trainees = get_trainees(form_id)
                # Convert to JSON format for the frontend
                form.trainees_data.data = json.dumps(existing_trainees)
            except Exception as e:
                logging.error(f"Error loading trainees for form {form_id}: {e}")
                form.trainees_data.data = "[]"

    if form.validate_on_submit():
        try:
            # Get trainees data from form
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                form.trainees_data.data = trainees_data

            # Get form data
            form_data = form.prepare_form_data()

            # Get the existing form to preserve the submitter
            existing_form = get_training_form(form_id)
            if existing_form and existing_form.get("submitter"):
                form_data["submitter"] = existing_form["submitter"]
            else:
                form_data["submitter"] = current_user.email

            # Validate required fields
            if not form_data.get("training_description"):
                flash("Training Description is required", "error")
                return render_template(
                    "index.html", form=form, edit_mode=True, form_id=form_id
                )

            # Update form data in database
            update_training_form(form_id, form_data)

            # Handle Attachments
            unique_folder = os.path.join(app.config["UPLOAD_FOLDER"], f"form_{form_id}")
            os.makedirs(unique_folder, exist_ok=True)

            # Process NEW attachments using SQLAlchemy ORM
            if "attachments" in request.files:
                new_files = request.files.getlist("attachments")
                descriptions = request.form.getlist("attachment_descriptions[]")
                
                if len(descriptions) < len(new_files):
                    descriptions.extend([""] * (len(new_files) - len(descriptions)))
                with db_session() as session:
                    for i, file in enumerate(new_files):
                        if file and file.filename:
                            filename = secure_filename(file.filename)
                            file_path = os.path.join(unique_folder, filename)
                            file.save(file_path)
                            description = (
                                descriptions[i] if i < len(descriptions) else ""
                            )
                            session.add(
                                Attachment(
                                    form_id=form_id,
                                    filename=filename,
                                    description=description,
                                )
                            )

            # Handle attachment DELETIONS using SQLAlchemy ORM
            delete_attachments = request.form.getlist("delete_attachments[]")
            if delete_attachments:
                with db_session() as session:
                    session.query(Attachment).filter(
                        Attachment.id.in_(delete_attachments),
                        Attachment.form_id == form_id,
                    ).delete(synchronize_session=False)

            # Handle attachment description UPDATES for existing files using SQLAlchemy ORM
            update_descriptions = request.form.getlist(
                "update_attachment_descriptions[]"
            )
            if update_descriptions:
                with db_session() as session:
                    for desc_json in update_descriptions:
                        try:
                            desc_data = json.loads(desc_json)
                            att_id = desc_data.get("id")
                            new_desc = desc_data.get("description", "")
                            if att_id:
                                att = (
                                    session.query(Attachment)
                                    .filter_by(id=att_id, form_id=form_id)
                                    .first()
                                )
                                if att:
                                    att.description = new_desc
                            else:
                                logging.warning(
                                    f"Skipping description update due to missing ID in JSON: {desc_json}"
                                )
                        except json.JSONDecodeError:
                            logging.error(
                                f"Invalid JSON in attachment description update: {desc_json}"
                            )
                        except KeyError:
                            logging.error(
                                f"Missing 'id' or 'description' key in JSON: {desc_json}"
                            )
            
            # Process travel expenses
            travel_expenses_data = request.form.get("travel_expenses_data")
            if travel_expenses_data:
                try:
                    from models import update_travel_expenses
                    travel_expenses = json.loads(travel_expenses_data)
                    if isinstance(travel_expenses, list):
                        update_travel_expenses(form_id, travel_expenses)
                        logging.info(f"Updated travel expenses for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logging.error(f"Error processing travel expenses: {e}")
                    # Don't fail the form update for travel expense errors
                    flash("Warning: There was an issue processing travel expenses, but the form was updated successfully.", "warning")
            
            # Process material expenses
            material_expenses_data = request.form.get("material_expenses_data")
            if material_expenses_data:
                try:
                    from models import update_material_expenses
                    material_expenses = json.loads(material_expenses_data)
                    if isinstance(material_expenses, list):
                        update_material_expenses(form_id, material_expenses)
                        logging.info(f"Updated material expenses for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logging.error(f"Error processing material expenses: {e}")
                    # Don't fail the form update for material expense errors
                    flash("Warning: There was an issue processing material expenses, but the form was updated successfully.", "warning")
            
            # Process trainees using the new table structure
            trainees_data = request.form.get("trainees_data")
            if trainees_data:
                try:
                    from models import update_trainees
                    trainees = json.loads(trainees_data)
                    if isinstance(trainees, list):
                        update_trainees(form_id, trainees)
                        logging.info(f"Updated trainees for form {form_id}")
                except (json.JSONDecodeError, Exception) as e:
                    logging.error(f"Error processing trainees: {e}")
                    # Don't fail the form update for trainee errors
                    flash("Warning: There was an issue processing trainees, but the form was updated successfully.", "warning")

            flash("Form updated successfully!", "success")
            return redirect(url_for("view_form", form_id=form_id))

        except Exception as e:
            logging.error(f"Error updating form: {str(e)}")
            flash(
                "An error occurred while updating the form. Please try again.", "danger"
            )

    # Load existing attachments to display in form using SQLAlchemy ORM
    with db_session() as session:
        existing_attachments = (
            session.query(Attachment).filter_by(form_id=form_id).all()
        )
        existing_attachments = [
            {
                "id": a.id,
                "form_id": a.form_id,
                "filename": a.filename,
                "description": a.description,
            }
            for a in existing_attachments
        ]

    # Load existing travel expenses
    existing_travel_expenses = []
    try:
        from models import get_travel_expenses
        existing_travel_expenses = get_travel_expenses(form_id)
    except Exception as e:
        logging.error(f"Error loading travel expenses for form {form_id}: {e}")

    # Load existing material expenses
    existing_material_expenses = []
    try:
        from models import get_material_expenses
        existing_material_expenses = get_material_expenses(form_id)
    except Exception as e:
        logging.error(f"Error loading material expenses for form {form_id}: {e}")

    return render_template(
        "index.html",
        form=form,
        edit_mode=True,
        form_id=form_id,
        existing_attachments=existing_attachments,
        existing_travel_expenses=existing_travel_expenses,
        existing_material_expenses=existing_material_expenses,
    )


@app.route("/api/employees")
@login_required
def get_employees():
    """API endpoint to get employees from the lookup module"""
    try:
        employees = get_lookup_data("employees")
        return jsonify(employees)
    except Exception as e:
        logging.error(f"Error fetching employee data via /api/employees: {str(e)}")
        return jsonify([])


@app.route("/api/lookup/<string:entity_type>")
@login_required
def api_lookup(entity_type):
    """Generic API endpoint to fetch lookup data for various entities."""
    logging.info(f"Received API lookup request for entity: {entity_type}")
    try:
        data = get_lookup_data(entity_type)
        if data is None: # get_lookup_data returns [] on error/not found, but explicit None check is safer
            logging.error(f"No data returned from get_lookup_data for entity: {entity_type}")
            return jsonify({"error": f"Data not found for {entity_type}"}), 404
        return jsonify(data)
    except Exception as e:
        logging.error(f"Error in /api/lookup/{entity_type} endpoint: {str(e)}", exc_info=True)
        return jsonify({"error": "Server error during lookup"}), 500


@app.route("/api/export_claim5_options")
@login_required
def export_claim5_options():
    if not is_admin_user(current_user):
        return jsonify({"error": "Unauthorized"}), 403
    from models import get_approved_forms_for_export
    import datetime

    forms = get_approved_forms_for_export()
    if not forms:
        return jsonify({"quarters": [], "min_date": None, "max_date": None})
    created_dates = [
        f.get("created_at") or f.get("submission_date") or f.get("start_date")
        for f in forms
        if f.get("created_at") or f.get("submission_date") or f.get("start_date")
    ]
    created_dates = [
        datetime.datetime.fromisoformat(str(d)[:10]) for d in created_dates
    ]
    min_date = min(created_dates).date().isoformat()
    max_date = max(created_dates).date().isoformat()

    quarters = sorted(
        {get_quarter(dt) for dt in created_dates}, key=lambda x: (int(x[1]), int(x[3:]))
    )
    return jsonify({"quarters": quarters, "min_date": min_date, "max_date": max_date})


@app.route("/export_claim5", methods=["GET", "POST"])
@login_required
def export_claim5():
    if not is_admin_user(current_user):
        flash("Unauthorized", "danger")
        return redirect(url_for("list_forms"))
    import json

    if request.method == "POST":
        data = request.get_json()
        selected_quarters = data.get("quarters", [])
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        from models import get_approved_forms_for_export
        import datetime

        forms = get_approved_forms_for_export()
        filtered_forms = []

        for f in forms:
            created = (
                f.get("created_at") or f.get("submission_date") or f.get("start_date")
            )
            if not created:
                continue
            dt = datetime.datetime.fromisoformat(str(created)[:10])
            quarter = get_quarter(dt)
            if selected_quarters and quarter in selected_quarters:
                filtered_forms.append(f)
            elif start_date and end_date:
                if start_date <= dt.date().isoformat() <= end_date:
                    filtered_forms.append(f)
        approved_forms = filtered_forms
    else:
        from models import get_approved_forms_for_export

        approved_forms = get_approved_forms_for_export()

    try:
        # Path to the new template Excel file
        template_path = os.path.join(
            "attached_assets", "Claim-Form-5-Training new GBER Rules.xlsx"
        )

        if not os.path.exists(template_path):
            flash("Template file not found.", "danger")
            return redirect(url_for("list_forms"))

        wb = load_workbook(template_path)
        
        # Get the sheets
        trainee_sheet = wb["Trainee"]
        external_trainer_sheet = wb["External Trainer"]
        internal_trainers_sheet = wb["Internal Trainers"]
        personnel_sheet = wb["Personnel Costs Lookup Table"]
        travel_sheet = wb["Travel"]
        materials_sheet = wb["Materials"]
        
        # Initialize personnel list to track unique employees
        personnel = []
        departments = []  # Track departments corresponding to personnel
        
        # Start row for data (header is on row 15)
        trainee_row = 16
        
        # Start row for external trainer data (header is on row 7)
        external_trainer_row = 8
        
        # Start row for internal trainers data (header is on row 8)
        internal_trainers_row = 9

        # Start row for travel data (header is on row 11)
        travel_row = 12

        # Start row for materials data (header is on row 13)
        materials_row = 14

        def process_trainee_sheet(form):
            nonlocal trainee_row
            try:
                # Get trainees data for this form from the new table structure
                trainees = []
                try:
                    from models import get_trainees
                    trainees = get_trainees(form['id'])
                except Exception as e:
                    logging.error(f"Error getting trainees for form {form['id']}: {e}")
                    return

                # If no trainees found, add a placeholder row
                if not trainees:
                    logging.warning(f"No trainees found for form {form['id']}")
                    return

                # Process each trainee
                for trainee in trainees:
                    # Extract email username (remove @ and everything after)
                    email = trainee.get("email", "")
                    trainee_name = email.split("@")[0] if "@" in email else email
                    
                    # Add trainee to personnel list if not already in
                    if trainee_name and trainee_name not in personnel:
                        personnel.append(trainee_name)
                        # Get department from the trainee record (stored in trainees table)
                        department = trainee.get("department", "")
                        departments.append(department)

                    # Fill the row with data according to requirements
                    trainee_sheet.cell(row=trainee_row, column=1).value = trainee_name  # Trainee Name
                    trainee_sheet.cell(row=trainee_row, column=2).value = form.get("training_name", "")  # Course Code/Name
                    
                    # Handle ida_class certification class extraction
                    ida_class = form.get("ida_class", "")
                    if ida_class == "Training not completed/ongoing":
                        certification_class = "Ongoing"
                    elif ida_class.startswith("Class "):
                        certification_class = ida_class[6:7]  # Extract letter (A, B, C, D)
                    else:
                        certification_class = ida_class
                    trainee_sheet.cell(row=trainee_row, column=3).value = certification_class  # Certification Class
                    
                    trainee_sheet.cell(row=trainee_row, column=5).value = form.get("training_hours", "")  # Training Hours
                    trainee_sheet.cell(row=trainee_row, column=8).value = form.get("start_date", "")  # Start Date
                    trainee_sheet.cell(row=trainee_row, column=9).value = form.get("end_date", "")  # End Date

                    # Handle trainer names based on training type
                    if form.get("training_type") == "Internal Training":
                        trainer_name = form.get("trainer_email", "").partition("@")[0]
                        # Add trainer to personnel list if not already in
                        if trainer_name and trainer_name not in personnel:
                            personnel.append(trainer_name)
                            # Use stored trainer department from training_forms table
                            trainer_department = form.get("trainer_department", "")
                            departments.append(trainer_department)
                        trainee_sheet.cell(row=trainee_row, column=10).value = trainer_name  # Internal Trainer Name
                        trainee_sheet.cell(row=trainee_row, column=11).value = ""  # External Trainer Name
                    else:
                        trainee_sheet.cell(row=trainee_row, column=10).value = ""  # Internal Trainer Name
                        trainee_sheet.cell(row=trainee_row, column=11).value = form.get("supplier_name", "")  # External Trainer Name

                    trainee_row += 1

            except Exception as e:
                logging.error(f"Error processing trainee sheet for form {form['id']}: {str(e)}", exc_info=True)

        def process_travel_expenses(form):
            nonlocal travel_row
            try:
                # Get travel expenses for this form
                from models import get_travel_expenses
                travel_expenses = get_travel_expenses(form['id'])
                
                if not travel_expenses:
                    return  # No travel expenses for this form
                
                # Process each travel expense
                for expense in travel_expenses:
                    # Column 1: Date
                    travel_sheet.cell(row=travel_row, column=1).value = expense.get("travel_date", "")
                    
                    # Column 2: Trainee/trainer name (extract username from email)
                    traveler_email = expense.get("traveler_email", "")
                    traveler_name = traveler_email.split("@")[0] if "@" in traveler_email else traveler_email
                    travel_sheet.cell(row=travel_row, column=2).value = traveler_name
                    
                    # Column 3: Travel type (columns C and D are merged, so write to C)
                    travel_mode = expense.get("travel_mode", "")
                    if travel_mode == "economy_flight":
                        travel_type = "Economy Flight"
                    elif travel_mode == "mileage":
                        travel_type = "Mileage"
                    elif travel_mode == "rail":
                        travel_type = "Rail"
                    elif travel_mode == "Bus":
                        travel_type = "bus"
                    else:
                        travel_type = travel_mode
                    travel_sheet.cell(row=travel_row, column=3).value = travel_type
                    
                    # Column 5: Travel cost (skip column 4 since C/D are merged)
                    cost = expense.get("cost", 0)
                    travel_sheet.cell(row=travel_row, column=5).value = cost if cost else 0
                    
                    # Column 6: Destination and course details
                    destination = expense.get("destination", "")
                    training_name = form.get("training_name", "")
                    destination_details = f"Destination: {destination}, Course Details: {training_name}"
                    travel_sheet.cell(row=travel_row, column=6).value = destination_details
                    
                    travel_row += 1
                    
            except Exception as e:
                logging.error(f"Error processing travel expenses for form {form['id']}: {str(e)}", exc_info=True)

        def process_material_expenses(form):
            nonlocal materials_row
            try:
                # Get material expenses for this form
                from models import get_material_expenses
                material_expenses = get_material_expenses(form['id'])
                
                if not material_expenses:
                    return  # No material expenses for this form
                
                # Process each material expense
                for expense in material_expenses:
                    # Column 1: Date
                    materials_sheet.cell(row=materials_row, column=1).value = expense.get("purchase_date", "")
                    
                    # Column 2: Supplier name
                    materials_sheet.cell(row=materials_row, column=2).value = expense.get("supplier_name", "")
                    
                    # Column 3: Invoice number (columns 3 and 4 are merged, so write to 3)
                    materials_sheet.cell(row=materials_row, column=3).value = expense.get("invoice_number", "")
                    
                    # Column 5: Course materials (material cost) - skip column 4 since 3/4 are merged
                    materials_sheet.cell(row=materials_row, column=5).value = expense.get("material_cost", 0)
                    
                    # Column 6: Course details (training name from the form)
                    materials_sheet.cell(row=materials_row, column=6).value = form.get("training_name", "")
                    
                    materials_row += 1
                    
            except Exception as e:
                logging.error(f"Error processing material expenses for form {form['id']}: {str(e)}", exc_info=True)

        # Process each approved form
        for form in approved_forms:
            process_trainee_sheet(form)
            process_travel_expenses(form)
            process_material_expenses(form)
            
            # Handle trainer sheets based on training type
            if form.get("training_type") == "Internal Training":
                # Add data to internal trainers sheet
                # Process internal trainer information
                try:
                    # Extract username from email (remove domain)
                    trainer_name = form.get("trainer_email", "").partition("@")[0]
                    # Add trainer to personnel list if not already in
                    if trainer_name and trainer_name not in personnel:
                        personnel.append(trainer_name)
                        # Use stored trainer department from training_forms table
                        trainer_department = form.get("trainer_department", "")
                        departments.append(trainer_department)
                    
                    internal_trainers_sheet.cell(row=internal_trainers_row, column=1).value = trainer_name
                    internal_trainers_sheet.cell(row=internal_trainers_row, column=3).value = form.get("training_name", "")
                    internal_trainers_sheet.cell(row=internal_trainers_row, column=4).value = form.get("training_hours", "")
                    
                    internal_trainers_row += 1
                    
                except Exception as e:
                    logging.error(f"Error processing internal trainers sheet for form {form['id']}: {str(e)}", exc_info=True)
            else:
                # Add data to external trainers sheet  
                # Process external trainer information
                try:
                    external_trainer_sheet.cell(row=external_trainer_row, column=1).value = form.get("start_date", "")
                    external_trainer_sheet.cell(row=external_trainer_row, column=2).value = form.get("supplier_name", "")
                    external_trainer_sheet.cell(row=external_trainer_row, column=3).value = form.get("invoice_number", "")
                    external_trainer_sheet.cell(row=external_trainer_row, column=4).value = form.get("training_name", "")
                    external_trainer_sheet.cell(row=external_trainer_row, column=5).value = form.get("course_cost", 0)
                    external_trainer_sheet.cell(row=external_trainer_row, column=6).value = form.get("training_description", "")
                    
                    external_trainer_row += 1
                    
                except Exception as e:
                    logging.error(f"Error processing external trainer sheet for form {form['id']}: {str(e)}", exc_info=True)

        # Populate Personnel Costs Lookup Table sheet
        # Header row starts at row 2, data starts at row 3
        personnel_row = 3
        for i, name in enumerate(personnel):
            personnel_sheet.cell(row=personnel_row, column=2).value = name  # Column B is name
            personnel_sheet.cell(row=personnel_row, column=3).value = departments[i]  # Column C is department
            personnel_row += 1

        # Create an in-memory file to store the Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logging.info(f"Exported {len(approved_forms)} approved forms to Excel template with {len(personnel)} unique personnel.")
        # Send the file to the user
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="claim5_export.xlsx",
        )

    except Exception as e:
        logging.error(f"Error exporting data to Excel template: {e}", exc_info=True)
        flash("An error occurred during the export process.", "danger")
        return redirect(url_for("list_forms"))


@app.errorhandler(500)
def internal_error(error):
    """Handle internal server errors"""
    flash("An internal server error occurred. Please try again later.", "danger")
    return render_template("index.html", form=TrainingForm(), now=datetime.now()), 500


@app.errorhandler(403)
def forbidden(e):
    # Flash a message for 403 errors and redirect to the previous page
    flash("You do not have permission to perform this action.", "danger")
    return redirect(request.referrer or url_for("index")), 403


# Add user information to layout template context
@app.context_processor
def inject_user():
    """Add user information to template context"""
    user_info = {}
    if current_user.is_authenticated:
        user_info = {
            "username": current_user.username,
            "display_name": current_user.display_name or current_user.username,
            "email": current_user.email,
            "first_name": current_user.first_name,
            "last_name": current_user.last_name,
            "is_admin": getattr(
                current_user, "is_admin", False
            ),  # Get admin status if available
        }
    return {"user_info": user_info}


@app.route("/my_submissions")
@login_required
def my_submissions():
    """Display a list of the current user's training form submissions"""
    form = SearchForm()
    search_term = request.args.get("search", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    training_type = request.args.get("training_type", "")
    approval_status = request.args.get("approval_status", "")
    delete_status = request.args.get("delete_status", "")
    sort_by = request.args.get("sort_by", "submission_date")
    sort_order = request.args.get("sort_order", "DESC")
    page = request.args.get("page", 1, type=int)

    # Populate form fields with current filter values
    form.search.data = search_term
    if date_from:
        try:
            form.date_from.data = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_to:
        try:
            form.date_to.data = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass
    form.training_type.data = training_type
    form.approval_status.data = approval_status
    form.delete_status.data = delete_status
    form.sort_by.data = sort_by
    form.sort_order.data = sort_order

    forms, total_count = get_user_training_forms(
        current_user.email,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        approval_status=approval_status,
        delete_status=delete_status,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
    )
    total_pages = (total_count + 9) // 10
    params = {
        "search": search_term,
        "date_from": date_from,
        "date_to": date_to,
        "training_type": training_type,
        "approval_status": approval_status,
        "delete_status": delete_status,
        "sort_by": sort_by,
        "sort_order": sort_order,
    }
    return render_template(
        "list.html",
        form=form,
        forms=forms,
        total_count=total_count,
        total_pages=total_pages,
        current_page=page,
        search_term=search_term,
        date_from=date_from,
        date_to=date_to,
        training_type=training_type,
        approval_status=approval_status,
        delete_status=delete_status,
        sort_by=sort_by,
        sort_order=sort_order,
        now=datetime.now(),
        params=params,
        has_filters=bool(search_term or date_from or date_to or training_type or approval_status or delete_status),
        total_forms=total_count,
        my_submissions=True,
        is_admin=is_admin_user(current_user),
    )


@app.route("/new")
@login_required
def new_form():
    """Display the training form"""
    logging.debug("=== NEW FORM ROUTE ===")
    form = TrainingForm()
    logging.debug(f"Form created with default data: {form.data}")
    logging.debug(f"Form errors (should be empty on initial load): {form.errors}")
    logging.debug(f"Location type field default value: {form.location_type.data}")
    return render_template("index.html", form=form, now=datetime.now())


@app.route("/leaderboard")
@login_required
def leaderboard():
    """Display the leaderboard of trainers by total training hours"""

    # Get all approved forms
    forms = get_approved_forms_for_export()
    training_hours = defaultdict(float)

    for form in forms:
        trainer_name = form.get("trainer_name")
        if not trainer_name:  # Skip forms without a trainer name
            continue

        try:
            from models import get_trainees
            trainees = get_trainees(form['id'])
            num_trainees = len(trainees) if trainees else 0
        except Exception as e:
            logging.error(f"Error getting trainees for leaderboard form {form['id']}: {e}")
            num_trainees = 0
        
        training_hours_val = float(form.get("training_hours") or 0)
        total_hours = training_hours_val * num_trainees
        training_hours[trainer_name] += total_hours

    # Sort trainers by total hours descending
    leaderboard_data = sorted(training_hours.items(), key=lambda x: x[1], reverse=True)
    names = [x[0] for x in leaderboard_data]
    hours = [x[1] for x in leaderboard_data]
    return render_template("leaderboard.html", names=names, hours=hours)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=app.config["DEBUG"])
